# coding=utf-8
import openpyxl


class HandExcel:
    # 加载excel
    def load_excel(self, excel_type):
        open_excel = openpyxl.load_workbook(excel_type)
        return open_excel

    # 根据sheet名获取对应的index
    def get_sheet_name(self, case_id, excel_type):
        sheet_name = self.load_excel(excel_type).sheetnames
        sheet_num = sheet_name.index(''.join([i for i in sheet_name if i.startswith(case_id[:3])]))
        # print('sheet_name-->', sheet_name)
        # print('sheet_num-->', sheet_num)
        return sheet_name, sheet_num

    # 加载所有sheet的内容
    def get_sheet_data(self, excel_type, index=None):
        sheet_name = self.load_excel(excel_type).sheetnames
        if index == None:
            index = 0
        data = self.load_excel(excel_type)[sheet_name[index]]
        return data

    # 获取某一个单元格内容
    def get_cell_value(self, case_id, row, cols, excel_type):
        sheet_num = self.get_sheet_name(case_id, excel_type)[1]
        data = self.get_sheet_data(excel_type, sheet_num).cell(row=row, column=cols).value
        return data

    # 获取行数
    def get_rows(self, excel_type, index=None):
        row = self.get_sheet_data(excel_type, index).max_row
        # print('row--->', row)
        return row

    # 获取某一行的内容
    def get_rows_value(self, row, excel_type, index=None):
        row_list = []
        for i in self.get_sheet_data(excel_type, index)[row]:
            row_list.append(i.value)
        return row_list

    # # 写入数据，仅支持单个sheet
    # def excel_write_data(self, row, cols, value, excel_type):
    #     wb = self.load_excel(excel_type)
    #     wr = wb.active
    #     wr.cell(row, cols, value)
    #     wb.save(excel_type)

    # 写入数据
    def excel_write_data(self, case_id, row, cols, value, excel_type):

        sheet_name, sheet_num = self.get_sheet_name(case_id, excel_type)
        wb = self.load_excel(excel_type)
        sheet1 = wb[sheet_name[sheet_num]]

        sheet1.cell(row, cols, value)
        wb.save(excel_type)

    # 获取某一列得数据
    def get_columns_value(self, excel_type, index=None, key=None):
        columns_list = []
        if key==None:
            key = 'A'
        columns_list_data = self.get_sheet_data(excel_type, index)[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    # 获取行号
    def get_rows_number(self, case_id, excel_type):
        num = 1
        sheet_num = self.get_sheet_name(case_id, excel_type)[1]
        cols_data = self.get_columns_value(excel_type, sheet_num)

        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num+1
        return num

    # 获取excel里面某个sheet的所有的数据
    def get_excel_data(self, excel_type, index=None):
        if index == None:
            index = 0
        data_list = []
        for i in range(self.get_rows(excel_type, index)-1):
            data_list.append(self.get_rows_value(i+2, excel_type, index))
        
        return data_list

    # 加载excel所有的内容
    def get_all_sheet(self, excel_type):
        sheet_name = self.load_excel(excel_type).sheetnames
        data_list = []

        for index in range(0, len(sheet_name)):
            data_list = data_list + (self.get_excel_data(excel_type, index))

        return data_list


excel_data = HandExcel()


if __name__ == "__main__":
    handle = HandExcel()
    print(handle.get_all_sheet('/Users/jiangyun/Documents/selenium/data/keywords.xlsx'))
